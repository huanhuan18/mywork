# -*- coding:utf-8 -*-
import openpyxl
import pprint
from openpyxl.utils import get_column_letter

wb1 = openpyxl.load_workbook(r"I:\disk_operation\2227\B11.xlsx")
sheet = wb1.active

alldata = {}

# fill in result_diff with each result and diff
for row in range(2, sheet.max_row+1):

    # each row in the spreadsheet has data
    casename = sheet["A" + str(row)].value
    result = sheet["B" + str(row)].value
    diff = sheet["D" + str(row)].value

    # make sure the key state exists
    alldata.setdefault(casename, {})   # if "result" exists,pass  else add key named "result"
    alldata[casename].setdefault("result", result)
    alldata[casename].setdefault("diff", diff)

    # print(alldata)

wb2 = openpyxl.load_workbook(r"I:\disk_operation\2227\B12.xlsx")
sheet = wb2.active

for row in range(2, sheet.max_row+1):
    casename = sheet["A" + str(row)].value
    result = sheet["B" + str(row)].value
    diff = sheet["D" + str(row)].value

    if result == "pass":
        continue
    else:
        if abs(diff - alldata[casename]["diff"]) < 0.0001:
            sheet["B" + str(row)] = "pass"
        else:
            continue
wb2.save(r"I:\disk_operation\2227\B12.xlsx")


# open a new text file and write the contents to it
resultFile = open("result_compare.py", "w")
resultFile.write("alldata = " + pprint.pformat(alldata))  # alldata transform to string

# other ways
# create a new excel
wb = openpyxl.Workbook()
sheet = wb.active

# change the name of the sheet
print(sheet.title)
sheet.title = "Happy2020"
print(wb.sheetnames)

# create new sheets
wb.create_sheet(index=0, title="First Sheet")
wb.create_sheet(index=1, title="Middle Sheet")
print(wb.sheetnames)

# delete sheet
wb.remove(wb["Middle Sheet"])
wb.save("excelcompare_bak.xlsx")

# write values to cell
sheet["A1"] = "Hello python"
print(sheet["A1"].value)

ws1 = wb.create_sheet("range names")
for row in range(1, 40):
    ws1.append(range(17))

ws2 = wb.create_sheet("List")
rows = [
    ["number", "Batch1", "Batch2"],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10]
]
for row in rows:
    ws2.append(row)

ws3 = wb.create_sheet(title="Data")
for row in range(5, 30):
    for col in range(15, 54):
        ws3.cell(column=col, row=row, value=get_column_letter(col))

wb.save(filename="empty_book.xlsx")